from flask import Flask, render_template, request, redirect, url_for, session, jsonify
import openpyxl
import os
import hashlib
from datetime import datetime

app = Flask(__name__)
app.secret_key = "miauroscopo_secret_key_2024"

EXCEL_FILE = "usuarios.xlsx"
ADMIN_USER = "admin"
ADMIN_PASS = "miauadmin"

HOROSCOPOS = {
    "Aries":       {"emoji": "🐯", "msg": "¡Hoy eres un gato salvaje lleno de energía! Lánzate a nuevas aventuras sin miedo. El universo felino te acompaña. Recuerda: los gatos siempre caen de pie. ¡Tú también lo harás!", "color": "#FF6B6B"},
    "Tauro":       {"emoji": "🌸", "msg": "Hoy el cosmos te invita a disfrutar tus placeres como un gato en el sol de la tarde. Busca comodidad y relax. Un momento de calma traerá grandes recompensas.", "color": "#FF8FAB"},
    "Géminis":     {"emoji": "🎭", "msg": "Tu curiosidad gatuna está al máximo hoy. Explora nuevas ideas, conversa y diviértete. Como el gato que persigue su propia cola, ¡el movimiento es tu aliado!", "color": "#FFD166"},
    "Cáncer":      {"emoji": "🏠", "msg": "Hoy el hogar es tu refugio, como un gatito que busca su lugar favorito. Rodéate de quienes amas. Las emociones fluyen como los ronroneos al anochecer.", "color": "#81B29A"},
    "Leo":         {"emoji": "👑", "msg": "¡Eres el rey o reina de la jungla hoy! Como el gato más majestuoso, brillas y todos te admiran. Acepta los halagos con elegancia felina. ¡Tu momento es ahora!", "color": "#F4A261"},
    "Virgo":       {"emoji": "🔍", "msg": "Tu instinto gatuno te guía hacia la perfección. Organiza, planifica y ejecuta con precisión de felino cazador. Los detalles son tu superpoder hoy.", "color": "#A8DADC"},
    "Libra":       {"emoji": "⚖️", "msg": "Busca el equilibrio como un gato caminando en la cornisa. La armonía en tus relaciones trae paz. Hoy es buen día para mediar y crear conexiones hermosas.", "color": "#C77DFF"},
    "Escorpio":    {"emoji": "🌙", "msg": "Tu misterio gatuno es irresistible hoy. Confía en tu intuición nocturna. Lo que se esconde en las sombras revelará verdades importantes. ¡Eres poderoso!", "color": "#6B4EFF"},
    "Sagitario":   {"emoji": "🚀", "msg": "¡Aventura felina al máximo! Hoy quieres explorar cada rincón del mundo. Como un gato que salta de techo en techo, ¡atrévete a ir más lejos!", "color": "#FF9F1C"},
    "Capricornio": {"emoji": "⛰️", "msg": "Con la paciencia de un gato cazando, alcanzarás tus metas hoy. La constancia es tu arma secreta. El éxito llega para quienes esperan con elegancia.", "color": "#2EC4B6"},
    "Acuario":     {"emoji": "⚡", "msg": "Tu espíritu gatuno independiente brilla hoy. Piensa diferente, innova y sorprende. Como el gato que hace lo que quiere, ¡sigue tu propio camino!", "color": "#4CC9F0"},
    "Piscis":      {"emoji": "🌊", "msg": "Tu sensibilidad felina está a flor de piel. Confía en tus sueños y déjate llevar por la corriente cósmica. Como el gato que observa el agua, ¡la magia está cerca!", "color": "#B5A0E6"},
}

SIGNOS = list(HOROSCOPOS.keys())

def hash_password(password):
    return hashlib.sha256(password.encode()).hexdigest()

def init_excel():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Usuarios"
        ws.append(["nombre", "telefono", "contrasena", "signo"])
        wb.save(EXCEL_FILE)

def load_users():
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    users = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            users.append({"nombre": row[0], "telefono": str(row[1]), "contrasena": row[2], "signo": row[3]})
    return users

def save_user(nombre, telefono, contrasena, signo):
    init_excel()
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append([nombre, telefono, hash_password(contrasena), signo])
    wb.save(EXCEL_FILE)

def user_exists(telefono):
    return any(u["telefono"] == telefono for u in load_users())

def find_user(telefono, contrasena):
    hashed = hash_password(contrasena)
    for u in load_users():
        if u["telefono"] == telefono and u["contrasena"] == hashed:
            return u
    return None

# ── Rutas ──────────────────────────────────────────────────────────────────────

@app.route("/")
def home():
    now = datetime.now().strftime("%A %d de %B de %Y")
    return render_template("home.html", now=now, signos=SIGNOS)

@app.route("/registro", methods=["GET", "POST"])
def registro():
    if request.method == "POST":
        nombre = request.form.get("nombre", "").strip()
        telefono = request.form.get("telefono", "").strip()
        contrasena = request.form.get("contrasena", "").strip()
        signo = request.form.get("signo", "").strip()
        if not nombre or not telefono or not contrasena or not signo:
            return render_template("registro.html", signos=SIGNOS, error="Por favor completa todos los campos.")
        if not telefono.startswith("+") or len(telefono) < 10:
            return render_template("registro.html", signos=SIGNOS, error="Ingresa tu teléfono en formato internacional. Ej: +521234567890")
        if user_exists(telefono):
            return render_template("registro.html", signos=SIGNOS, error="Este número ya está registrado.")
        save_user(nombre, telefono, contrasena, signo)
        return render_template("registro.html", signos=SIGNOS, success=nombre, signo=signo)
    return render_template("registro.html", signos=SIGNOS)

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        telefono = request.form.get("telefono", "").strip()
        contrasena = request.form.get("contrasena", "").strip()
        user = find_user(telefono, contrasena)
        if user:
            session["user"] = user
            return redirect(url_for("horoscopo"))
        return render_template("login.html", error="Número o contraseña incorrectos. ¡Intenta de nuevo, gatito!")
    return render_template("login.html")

@app.route("/horoscopo")
def horoscopo():
    user = session.get("user")
    if not user:
        return redirect(url_for("login"))
    signo = user["signo"]
    horo = HOROSCOPOS.get(signo, {"emoji": "🐱", "msg": "¡El cosmos felino tiene algo especial para ti!", "color": "#FF85A1"})
    today = datetime.now().strftime("%d de %B de %Y")
    return render_template("horoscopo.html", user=user, horo=horo, signo=signo, today=today)

@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("home"))

@app.route("/admin", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        u = request.form.get("usuario", "").strip()
        p = request.form.get("contrasena", "").strip()
        if u == ADMIN_USER and p == ADMIN_PASS:
            session["admin"] = True
            return redirect(url_for("admin_panel"))
        return render_template("admin_login.html", error="Credenciales incorrectas. ¡El gato guardián no te deja pasar!")
    return render_template("admin_login.html")

@app.route("/admin/panel")
def admin_panel():
    if not session.get("admin"):
        return redirect(url_for("admin_login"))
    users = load_users()
    signos_count = {}
    for u in users:
        signos_count[u["signo"]] = signos_count.get(u["signo"], 0) + 1
    return render_template("admin_panel.html", users=users, signos_count=signos_count, total=len(users))

@app.route("/admin/enviar", methods=["POST"])
def admin_enviar():
    if not session.get("admin"):
        return jsonify({"ok": False, "msg": "No autorizado"})
    try:
        import pywhatkit
    except ImportError:
        return jsonify({"ok": False, "msg": "pywhatkit no está instalado. Ejecuta: pip install pywhatkit"})

    users = load_users()
    if not users:
        return jsonify({"ok": False, "msg": "No hay usuarios registrados."})

    now = datetime.now()
    send_hour = now.hour
    send_min = now.minute + 2
    if send_min >= 60:
        send_min = 0
        send_hour += 1

    errors = []
    sent = 0
    for user in users:
        try:
            signo = user["signo"]
            horo = HOROSCOPOS.get(signo, {})
            mensaje = (
                f"🐱✨ *MIAUROSCOPO DEL DÍA* ✨🐱\n\n"
                f"¡Hola {user['nombre']}! 😺\n"
                f"⭐ *Tu signo: {signo}* {horo.get('emoji','')}\n\n"
                f"{horo.get('msg','')}\n\n"
                f"📅 {now.strftime('%d/%m/%Y')}\n"
                f"_Con amor felino, Miauroscopo_ 🐾"
            )
            pywhatkit.sendwhatmsg(user["telefono"], mensaje, send_hour, send_min, wait_time=20, tab_close=True)
            sent += 1
            send_min += 3
            if send_min >= 60:
                send_min = 0
                send_hour += 1
        except Exception as e:
            errors.append(f"{user['nombre']}: {str(e)}")

    if errors:
        return jsonify({"ok": True, "msg": f"Enviados {sent}/{len(users)}. Errores: {', '.join(errors[:3])}"})
    return jsonify({"ok": True, "msg": f"¡{sent} horóscopo(s) enviado(s) con éxito! 🎉"})

@app.route("/admin/logout")
def admin_logout():
    session.pop("admin", None)
    return redirect(url_for("home"))

if __name__ == "__main__":
    init_excel()
    app.run(debug=True)
